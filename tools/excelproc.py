from openpyxl import Workbook
from openpyxl import load_workbook

def procQL_REQUEST():
    wb = load_workbook("QL_REQUEST_template.xlsx")
    ws = wb.get_sheet_by_name("QUICKLOOK")
    ws["B2"] = "Installer_DEV_D2_CL375477\n+MGC_DEV_D2_CL333333\n+FEMPlguin_local_build"
    wb.save("QL_REQUEST.xlsx")

if "__main__"==__name__:
    procQL_REQUEST()